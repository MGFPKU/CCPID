#!/usr/bin/env python3
"""One-off recolor of the 概览/Summary template sheets to the CCPID webpage
module palette (2026-09-01):

- Header text  #44749F (steel blue)  -> #0E5C46 (module dark green).
- Light borders #DEE0E3              -> #E2E8E5 (module border).

Backs up each template to *.bak_recolor before rewriting.
(The blue->green data-bar swap lives permanently in export_workbooks.py:
DATABAR_COLOR is now 30% #14936f.)
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Color

from export_workbooks import LANG_CONFIG, ROOT

HEADER_FROM, HEADER_TO = "FF44749F", "FF0E5C46"
BORDER_FROM, BORDER_TO = "FFDEE0E3", "FFE2E8E5"


def recolor(ws) -> int:
    changed = 0
    for row in ws.iter_rows():
        for cell in row:
            font = cell.font
            if font and font.color and font.color.rgb == HEADER_FROM:
                cell.font = font.copy(color=Color(rgb=HEADER_TO))
                changed += 1
            border = cell.border
            for side_name in ("left", "right", "top", "bottom"):
                side = getattr(border, side_name)
                if side and side.color and side.color.rgb == BORDER_FROM:
                    side.color = Color(rgb=BORDER_TO)
                    changed += 1
    return changed


def main() -> int:
    total = 0
    for lang, config in LANG_CONFIG.items():
        path: Path = config["template"]
        wb = load_workbook(path)
        ws = wb[config["summary_sheet"]]
        changed = recolor(ws)
        if not changed:
            print(f"{path.name}: no changes")
            continue
        shutil.copy2(path, str(path) + ".bak_recolor")
        wb.save(path)
        print(
            f"{path.name} [{ws.title}]: {changed} cell(s) recolored "
            f"(backup: {path.name}.bak_recolor)"
        )
        total += changed
    print(f"\nTotal cells changed: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
