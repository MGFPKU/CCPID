#!/usr/bin/env python3
"""One-off: copy the user-added reference sheets from outputs/CCPID_en.xlsx
into inputs/template_en.xlsx so the export pipeline preserves them.

Sheets copied: Cover, Classification framework, Categories and groups,
Attributes definitions, Ended.
"""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "outputs" / "CCPID_en.xlsx"
DST = ROOT / "inputs" / "template_en.xlsx"
SHEETS = [
    "Cover",
    "Classification framework",
    "Categories and groups",
    "Attributes definitions",
    "Ended",
]


def copy_sheet(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.protection = copy(cell.protection)
                new_cell.number_format = cell.number_format

    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))

    for key, dim in src_ws.column_dimensions.items():
        letter = get_column_letter(key) if isinstance(key, int) else key
        new_dim = dst_ws.column_dimensions[letter]
        new_dim.width = dim.width
        new_dim.hidden = dim.hidden
        new_dim.bestFit = dim.bestFit

    for key, dim in src_ws.row_dimensions.items():
        new_dim = dst_ws.row_dimensions[key]
        new_dim.height = dim.height
        new_dim.hidden = dim.hidden

    if src_ws.sheet_properties.tabColor is not None:
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor


def main() -> None:
    src_wb = load_workbook(SRC)
    dst_wb = load_workbook(DST)

    for name in SHEETS:
        if name not in src_wb.sheetnames:
            raise RuntimeError(f"Sheet '{name}' not found in {SRC}")
        if name in dst_wb.sheetnames:
            raise RuntimeError(f"Sheet '{name}' already exists in {DST}")
        copy_sheet(src_wb[name], dst_wb.create_sheet(name))
        print(f"Copied '{name}' ({src_wb[name].max_row} rows)")

    dst_wb.save(DST)
    print(f"Saved {DST}")


if __name__ == "__main__":
    main()
