#!/usr/bin/env python3
"""One-off: add CN versions of the EN reference sheets to inputs/template_cn.xlsx.

- Copy Cover/Classification framework/Categories and groups/Attributes
  definitions from the EN template with Chinese tab names (content identical
  to the English version).
- Create the 已终止 (Ended) sheet with Chinese headers and the 7 curated
  historical rows mapped from the EN template's Ended sheet.
"""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
EN_TEMPLATE = ROOT / "inputs" / "template_en.xlsx"
CN_TEMPLATE = ROOT / "inputs" / "template_cn.xlsx"

SHEET_RENAMES = {
    "Cover": "封面",
    "Classification framework": "分类框架",
    "Categories and groups": "类别与组别",
    "Attributes definitions": "属性定义",
}

ENDED_CN_HEADERS = [
    "英文名",
    "本国工具名称",
    "状态",
    "开始年份",
    "结束年份",
    "工具类别",
    "工具组别",
    "排放部门",
    "子行业",
    "法律文件名称",
    "其他网页链接",
]

CATEGORY_CN = {"Economic": "经济工具"}
GROUP_CN = {"Subsidy": "补贴", "Tax": "税收"}
SECTOR_CN = {"Transport": "交通", "Energy": "能源"}
SUBSECTOR_CN = {
    "Vehicle": "车辆",
    "Solar power; wind power": "太阳能发电；风电",
    "Wind power": "风电",
    "Nuclear power": "核电",
    "Coal": "煤炭",
}


def year_only(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if "/" in text:
        return text.split("/")[-1]
    return text[:4]


def copy_sheet(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.protection = copy(cell.protection)
                new_cell.number_format = cell.number_format

    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))

    for key, dim in src_ws.column_dimensions.items():
        letter = get_column_letter(key) if isinstance(key, int) else key
        new_dim = dst_ws.column_dimensions[letter]
        new_dim.width = dim.width
        new_dim.hidden = dim.hidden
        new_dim.bestFit = dim.bestFit

    for key, dim in src_ws.row_dimensions.items():
        new_dim = dst_ws.row_dimensions[key]
        new_dim.height = dim.height
        new_dim.hidden = dim.hidden


def build_ended_sheet(en_wb, cn_wb) -> None:
    src = en_wb["Ended"]
    dst = cn_wb.create_sheet("已终止")
    dst.append(ENDED_CN_HEADERS)
    for cell in dst[1]:
        cell.font = Font(bold=True)

    curated = []
    for row in src.iter_rows(min_row=2, values_only=True):
        values = list(row[:11])
        if not any(v not in (None, "") for v in values):
            continue
        curated.append(values)

    for english_name, domestic_name, _status, start, end, category, group, sector, subsector, _legal, _links in curated:
        dst.append(
            [
                english_name,
                domestic_name,
                "已终止",
                year_only(start),
                year_only(end),
                CATEGORY_CN.get(str(category or "").strip(), str(category or "")),
                GROUP_CN.get(str(group or "").strip(), str(group or "")),
                SECTOR_CN.get(str(sector or "").strip(), str(sector or "")),
                SUBSECTOR_CN.get(str(subsector or "").strip(), str(subsector or "")),
                "",
                "",
            ]
        )
    for col_idx, width in enumerate([46, 30, 10, 12, 12, 12, 12, 14, 20, 34, 34], start=1):
        dst.column_dimensions[get_column_letter(col_idx)].width = width
    dst.freeze_panes = "A2"
    print(f"Created 已终止 with {len(curated)} curated rows")


def main() -> None:
    en_wb = load_workbook(EN_TEMPLATE)
    cn_wb = load_workbook(CN_TEMPLATE)

    for en_name, cn_name in SHEET_RENAMES.items():
        if cn_name in cn_wb.sheetnames:
            raise RuntimeError(f"Sheet '{cn_name}' already exists in {CN_TEMPLATE}")
        copy_sheet(en_wb[en_name], cn_wb.create_sheet(cn_name))
        print(f"Copied '{en_name}' -> '{cn_name}'")

    if "已终止" in cn_wb.sheetnames:
        raise RuntimeError("Sheet '已终止' already exists in CN template")
    build_ended_sheet(en_wb, cn_wb)

    cn_wb.save(CN_TEMPLATE)
    print(f"Saved {CN_TEMPLATE}")


if __name__ == "__main__":
    main()
