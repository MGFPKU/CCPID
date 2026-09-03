#!/usr/bin/env python3
"""Export working CSV data into reviewable Excel workbooks.

The CSV files stay useful for scripted updates and diffs. This script produces
the project deliverables from the official templates:

- outputs/CCPID_cn.xlsx
- outputs/CCPID_en.xlsx, when English CSVs exist
- outputs/evidence_log.xlsx
"""

from __future__ import annotations

import argparse
import csv
import datetime
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from policy_id import id_code_lookup, load_id_codes


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUTS_DIR = ROOT / "outputs"
CLASSIFICATION_DIR = ROOT / "inputs" / "classification"

LANG_CONFIG = {
    "cn": {
        "template": ROOT / "inputs" / "template_cn.xlsx",
        "output": "CCPID_cn.xlsx",
        "approaches_sheet": "路径汇总",
        "overview_sheet": "工具总览",
        "summary_sheet": "概览",
        "cover_sheet": "封面",
        "approaches_headers": ["类别", "组别", "路径", "缩写", "定义", "排放部门", "减缓相关性", "作用渠道", "工具数量"],
        "sheets": {
            "economic_instruments": "经济工具",
            "regulatory_instruments": "规制工具",
            "government_i_c": "政府投资与消费",
            "information_instruments": "信息工具",
            "voluntary_approaches": "自愿措施",
        },
        "front_sheets": ["封面", "分类框架", "类别与组别"],
        "back_sheets": ["属性定义", "已终止"],
        "ended_sheet": "已终止",
    },
    "en": {
        "template": ROOT / "inputs" / "template_en.xlsx",
        "output": "CCPID_en.xlsx",
        "approaches_sheet": "Approaches",
        "overview_sheet": "Instruments overview",
        "summary_sheet": "Summary",
        "cover_sheet": "Cover",
        "approaches_headers": [
            "Category",
            "Group",
            "Approach",
            "Abbreviation",
            "Definition",
            "Emission sector",
            "Mitigation relevance",
            "Functioning channel",
            "Number of instruments",
        ],
        "sheets": {
            "economic_instruments": "Economic instruments",
            "regulatory_instruments": "Regulatory instruments",
            "government_i_c": "Government I&C",
            "information_instruments": "Information instruments",
            "voluntary_approaches": "Voluntary approaches",
        },
        # Static reference sheets stored in the template (never rewritten from CSVs)
        "front_sheets": ["Cover", "Classification framework", "Categories and groups"],
        "back_sheets": ["Attributes definitions", "Ended"],
        "ended_sheet": "Ended",
    },
}

# Module accent #14936f at 30% alpha — matches the webpage data bars
# (rgba(20,147,111,0.30) over white ≈ #B9DED3).
DATABAR_COLOR = "4D14936F"

SUMMARY_DATABAR_RANGES = {
    "cn": [
        "C10:D25",
        "I10:J14",
        "I18:J24",
        "I29:J31",
        "I35:J36",
        "I40:N46",
    ],
    "en": [
        "C10:D25",
        "I10:J14",
        "I18:J24",
        "I29:J31",
        "I35:J36",
        "I40:N46",
    ],
}

X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
XM_NS = "http://schemas.microsoft.com/office/excel/2006/main"
X14_CF_RULE_URI = "{B025F937-C7B1-47D3-B67F-A62EFF666E3E}"
X14_CFMTS_URI = "{78C0D931-6437-407d-A8EE-F0AAD7539E65}"

APPROACH_ABBREVIATIONS = {
    "emissions trading system": "ETS",
    "tradable energy efficiency obligations or certificates": "EEO",
    "tradable renewable electricity credits, quota or tradable performance standards": "REC",
    "tradable performance standards": "TPS",
}

APPROACH_GROUP_SORT_ORDER = {
    # Economic instruments: Trading scheme, Subsidy, Tax, Administered price
    "trading scheme": 10,
    "交易机制": 10,
    "subsidy": 20,
    "补贴": 20,
    "tax": 30,
    "税收": 30,
    "administered price": 40,
    "行政定价": 40,
    # Regulatory instruments: Technology standard, Performance standard, Framework regulation
    "technology standard": 50,
    "技术标准": 50,
    "performance standard": 51,
    "绩效标准": 51,
    "framework regulation": 52,
    "框架性规制": 52,
    # Government I&C: Public investment, Public procurement, Public appraisal rules
    "public investment": 70,
    "公共投资": 70,
    "public procurement": 71,
    "公共采购": 71,
    "public appraisal rules": 72,
    "公共评估规则": 72,
    # Information instruments: Comparative energy efficiency label, Reporting requirements, Capacity building and public awareness
    "comparative energy efficiency label": 90,
    "比较性能效标签": 90,
    "reporting requirements": 91,
    "报告与披露要求": 91,
    "capacity building and public awareness": 92,
    "能力建设与公众意识": 92,
    # Voluntary approaches
    "voluntary trading system": 110,
    "自愿交易体系": 110,
    "voluntary information instrument": 111,
    "自愿性信息工具": 111,
    "voluntary target": 112,
    "自愿性目标": 112,
}


def normalise_key(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def strip_parenthetical(value: str) -> str:
    return re.sub(r"[（(][^）)]*[）)]", "", value).strip()


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def split_multi(value: str) -> list[str]:
    return [part.strip() for part in value.replace("；", ";").split(";") if part.strip()]


def append_unique(target: list[str], value: str) -> None:
    for part in split_multi(value):
        if part not in target:
            target.append(part)


def approach_abbreviation(approach: str) -> str:
    _, schema_approach_codes = load_id_codes()
    return id_code_lookup(approach, schema_approach_codes) or APPROACH_ABBREVIATIONS.get(approach.strip().casefold(), "")


def get_field(row: dict[str, str], headers: list[str], *names: str) -> str:
    by_norm = {normalise_key(header): header for header in headers}
    for name in names:
        source_header = by_norm.get(normalise_key(name))
        if source_header:
            return row.get(source_header, "")
    return ""


def markdown_table_rows(path: Path) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line.startswith("|") or not line.endswith("|"):
            continue
        if set(line.replace("|", "").replace(":", "").replace("-", "").strip()) <= set():
            continue
        rows.append([cell.strip() for cell in line.strip("|").split("|")])
    return rows


def load_english_approaches() -> dict[str, dict[str, str]]:
    rows = markdown_table_rows(CLASSIFICATION_DIR / "approaches_en.md")
    if not rows:
        return {}
    headers = rows[0]
    approaches: dict[str, dict[str, str]] = {}
    for cells in rows[1:]:
        if len(cells) != len(headers):
            continue
        row = dict(zip(headers, cells))
        approach = row.get("Approach", "")
        if approach:
            row["_order"] = str(len(approaches))
            approaches[normalise_key(approach)] = row
            approaches[normalise_key(strip_parenthetical(approach))] = row

    # Cross-reference: map Chinese approach names to matching English entries
    cn_approaches = load_chinese_approaches()
    seen_ids: set[int] = set()
    for cn_entry in cn_approaches.values():
        entry_id = id(cn_entry)
        if entry_id in seen_ids:
            continue
        seen_ids.add(entry_id)
        en_name = cn_entry.get("路径_en", "")
        cn_name = cn_entry.get("路径", "")
        if en_name and cn_name:
            en_key = normalise_key(en_name)
            cn_key = normalise_key(cn_name)
            if en_key in approaches and cn_key not in approaches:
                approaches[cn_key] = approaches[en_key]

    # Enrich entries with abbreviations from the abbreviation tables
    for raw_line in (CLASSIFICATION_DIR / "approaches_en.md").read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line.startswith("|") or not line.endswith("|"):
            continue
        if set(line.replace("|", "").replace(":", "").replace("-", "").strip()) <= set():
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) == 2 and cells[0] != "Approach":
            key = normalise_key(cells[0])
            if key in approaches:
                approaches[key]["Abbreviation"] = cells[1]

    return approaches


def heading_chinese_name(heading: str) -> str:
    parts = heading.strip().split()
    return parts[-1] if parts else heading.strip()


def load_chinese_approaches() -> dict[str, dict[str, str]]:
    path = CLASSIFICATION_DIR / "approaches_cn.md"
    current_category = ""
    current_group = ""
    in_approach_table = False
    approaches: dict[str, dict[str, str]] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if line.startswith("## "):
            current_category = heading_chinese_name(line.removeprefix("## "))
            in_approach_table = False
            continue
        if line.startswith("### "):
            current_group = heading_chinese_name(line.removeprefix("### "))
            in_approach_table = False
            continue
        if not line.startswith("|") or not line.endswith("|") or line.startswith("|---"):
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) >= 3 and cells[0] == "Approach" and cells[2] == "中文定义":
            in_approach_table = True
            continue
        if not in_approach_table or len(cells) < 3 or cells[0] == "Approach":
            continue
        chinese_name = cells[1]
        order = len(approaches)
        approaches[normalise_key(chinese_name)] = {
            "类别": current_category,
            "组别": current_group,
            "路径": chinese_name,
            "路径_en": cells[0],
            "缩写": approach_abbreviation(cells[0]),
            "定义": cells[2],
        }
        approaches[normalise_key(chinese_name)]["_order"] = str(order)
        approaches[normalise_key(strip_parenthetical(chinese_name))] = approaches[normalise_key(chinese_name)]
        approaches[normalise_key(cells[0])] = approaches[normalise_key(chinese_name)]
    return approaches


def clear_existing_data(ws, header_row: int = 2) -> None:
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)


def sheet_headers(ws, header_row: int = 2) -> list[str]:
    return [cell.value or "" for cell in ws[header_row]]


def style_output_row(ws, row_index: int, headers: list[str]) -> None:
    row_type_column = None
    for index, header in enumerate(headers, start=1):
        if normalise_key(str(header)) == normalise_key("Instrument / subscheme") or str(header) == "工具/子方案":
            row_type_column = index
            break
    if row_type_column is None:
        return

    row_type = str(ws.cell(row=row_index, column=row_type_column).value or "").strip().casefold()
    is_instrument = row_type in {"instrument", "工具"}
    for cell in ws[row_index]:
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            bold=is_instrument,
            italic=cell.font.italic,
            vertAlign=cell.font.vertAlign,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color=cell.font.color,
        )


def write_rows(ws, csv_headers: list[str], rows: list[dict[str, str]], header_row: int = 2) -> None:
    headers = sheet_headers(ws, header_row)
    by_norm = {normalise_key(header): header for header in csv_headers}
    clear_existing_data(ws, header_row)
    for row_index, source in enumerate(rows, start=header_row + 1):
        for col_index, header in enumerate(headers, start=1):
            source_header = by_norm.get(normalise_key(str(header)))
            ws.cell(row=row_index, column=col_index, value=source.get(source_header, "") if source_header else "")
        style_output_row(ws, row_index, headers)
    _apply_status_formulas(ws, header_row, len(rows))


def _apply_status_formulas(ws, header_row: int, row_count: int) -> None:
    """Replace static Status cell values with date-based formulas.

    Formula logic:
      - Start date empty/N/A  =>  In force/生效 (no temporal boundary)
      - TODAY < Start date     =>  Scheduled/计划实施
      - End date empty/N/A
        or TODAY < End date    =>  In force/生效
      - Otherwise               =>  Ended/已终止
    """
    headers = [str(c.value or "") for c in ws[header_row]]
    try:
        start_col_idx = next(i for i, h in enumerate(headers) if h in ("Start date", "生效日期"))
        end_col_idx = next(i for i, h in enumerate(headers) if h in ("End date", "终止日期"))
        status_col_idx = next(i for i, h in enumerate(headers) if h in ("Status", "状态"))
    except StopIteration:
        return  # one of the required columns is missing — skip

    from openpyxl.utils import get_column_letter
    r_col = get_column_letter(start_col_idx + 1)
    s_col = get_column_letter(end_col_idx + 1)

    is_cn = "状态" in str(ws.cell(row=header_row, column=status_col_idx + 1).value or "")
    if is_cn:
        formula_tmpl = (
            '=IF({r}{row}="","生效",'
            'IF({r}{row}="N/A","生效",'
            'IF(TODAY()<DATE(VALUE(RIGHT({r}{row},4)),VALUE(MID({r}{row},4,2)),VALUE(LEFT({r}{row},2))),"计划实施",'
            'IF({s}{row}="","生效",'
            'IF({s}{row}="N/A","生效",'
            'IF(TODAY()<DATE(VALUE(RIGHT({s}{row},4)),VALUE(MID({s}{row},4,2)),VALUE(LEFT({s}{row},2))),"生效","已终止"))))))'
        )
    else:
        formula_tmpl = (
            '=IF({r}{row}="","In force",'
            'IF({r}{row}="N/A","In force",'
            'IF(TODAY()<DATE(VALUE(RIGHT({r}{row},4)),VALUE(MID({r}{row},4,2)),VALUE(LEFT({r}{row},2))),"Scheduled",'
            'IF({s}{row}="","In force",'
            'IF({s}{row}="N/A","In force",'
            'IF(TODAY()<DATE(VALUE(RIGHT({s}{row},4)),VALUE(MID({s}{row},4,2)),VALUE(LEFT({s}{row},2))),"In force","Ended"))))))'
        )

    for row_num in range(header_row + 1, header_row + row_count + 1):
        formula = formula_tmpl.format(r=r_col, s=s_col, row=row_num)
        ws.cell(row=row_num, column=status_col_idx + 1).value = formula


def write_instruments_overview_sheet(wb, lang: str) -> None:
    """Populate the Instruments overview sheet with live cell references.

    Only runs when the template contains an "Instruments overview" sheet.
    Writes Excel cross-sheet reference formulas so that every overview cell
    updates automatically when the source sheet value changes.
    """
    config = LANG_CONFIG[lang]
    sheet_name = config.get("overview_sheet")
    if not sheet_name or sheet_name not in wb.sheetnames:
        return

    ov_ws = wb[sheet_name]

    # Clear existing data rows (row 3+) while preserving headers
    if ov_ws.max_row > 2:
        ov_ws.delete_rows(3, ov_ws.max_row - 2)

    from openpyxl.utils import get_column_letter

    # Map overview column -> source sheet column letter via 1-based index
    src_col_letter = {
        1: get_column_letter(1),   # Policy Instrument ID
        2: get_column_letter(4),   # Approach
        3: get_column_letter(8),   # English instrument name
        4: get_column_letter(7),   # Domestic instrument name
        5: get_column_letter(22),  # Status
        6: get_column_letter(18),  # Start date
        7: get_column_letter(19),  # End date
        8: get_column_letter(12),  # Mitigation relevance
        10: get_column_letter(3),  # Group
        11: get_column_letter(5),  # Emission sector
        12: get_column_letter(6),  # Sub-sector
        13: get_column_letter(13), # Functioning channel
    }

    row_idx = 3
    total = 0
    for sheet_label in config["sheets"].values():
        if sheet_label not in wb.sheetnames:
            continue
        src_ws = wb[sheet_label]
        # Quote sheet name for formula safety (handles spaces and special chars)
        quoted_sheet = f"'{sheet_label}'"
        for src_row in range(3, src_ws.max_row + 1):
            pid = str(src_ws.cell(row=src_row, column=1).value or "").strip()
            if not pid or pid == "None":
                continue
            # Only include instruments, not subschemes
            inst_type = str(src_ws.cell(row=src_row, column=2).value or "").strip()
            if inst_type in ("Subscheme", "子方案"):
                continue

            for ov_col, src_letter in src_col_letter.items():
                formula = f"={quoted_sheet}!{src_letter}{src_row}"
                ov_ws.cell(row=row_idx, column=ov_col, value=formula)

            # Category (col 9) is static — the sheet name itself
            ov_ws.cell(row=row_idx, column=9, value=sheet_label)
            row_idx += 1
            total += 1

    # Position between Summary (index 0) and first data sheet
    target_idx = 2
    current_idx = wb.sheetnames.index(sheet_name)
    if current_idx != target_idx:
        wb._sheets.insert(target_idx, wb._sheets.pop(current_idx))

    # Apply alignment for data rows
    for r in range(3, ov_ws.max_row + 1):
        for c in range(1, 14):
            ov_ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")


SUMMARY_LAYOUT = {
    "cn": {
        "groups": [
            (10, "交易机制"), (11, "补贴"), (12, "税收"), (13, "行政定价"),
            (14, "技术标准"), (15, "绩效标准"), (16, "框架性规制"),
            (17, "公共采购"), (18, "公共投资"), (19, "公共评估规则"),
            (20, "比较性能效标签"), (21, "报告与披露要求"), (22, "能力建设与公众意识"),
            (23, "自愿性信息工具"), (24, "自愿性目标"), (25, "自愿交易体系"),
        ],
        "cats": [
            (10, "经济工具"), (11, "规制工具"), (12, "政府投资与消费"),
            (13, "信息工具"), (14, "自愿措施"),
        ],
        "ov_cats": [
            (10, "经济工具"), (11, "规制工具"), (12, "政府投资与消费"),
            (13, "信息工具"), (14, "自愿措施"),
        ],
        "sectors": [
            (18, "能源"), (19, "工业"), (20, "建筑"), (21, "交通"),
            (22, "农业、林业和其他土地利用"), (23, "废弃物"), (24, "跨部门"),
        ],
        "channels": [(29, "供给侧"), (30, "需求侧"), (31, "环境")],
        "mitigations": [(35, "直接"), (36, "间接")],
        "approaches_label": "路径数量",
        "instruments_label": "工具数量",
        "cross_sectors": [
            (40, "能源"), (41, "工业"), (42, "建筑"), (43, "交通"),
            (44, "农业、林业和其他土地利用"), (45, "废弃物"), (46, "跨部门"),
        ],
        "cross_cats": [
            ("I", "经济工具"), ("J", "规制工具"), ("K", "政府投资与消费"),
            ("L", "信息工具"), ("M", "自愿措施"),
        ],
    },
    "en": {
        "groups": [
            (10, "Trading scheme"), (11, "Subsidy"), (12, "Tax"), (13, "Administered price"),
            (14, "Technology standard"), (15, "Performance standard"), (16, "Framework regulation"),
            (17, "Public procurement"), (18, "Public investment"), (19, "Public appraisal rules"),
            (20, "Comparative energy efficiency label"), (21, "Reporting requirements"),
            (22, "Capacity building and public awareness"),
            (23, "Voluntary information instrument"), (24, "Voluntary target"), (25, "Voluntary trading system"),
        ],
        "cats": [
            (10, "Economic"), (11, "Regulatory"), (12, "Government investment and consumption"),
            (13, "Information"), (14, "Voluntary"),
        ],
        "ov_cats": [
            (10, "Economic instruments"), (11, "Regulatory instruments"), (12, "Government I&C"),
            (13, "Information instruments"), (14, "Voluntary approaches"),
        ],
        "sectors": [
            (18, "Energy"), (19, "Industry"), (20, "Buildings"), (21, "Transport"),
            (22, "AFOLU"), (23, "Waste"), (24, "Cross-sectoral"),
        ],
        "channels": [(29, "Supply-side"), (30, "Demand-side"), (31, "Environment")],
        "mitigations": [(35, "Direct"), (36, "Indirect")],
        "approaches_label": "Number of Approaches",
        "instruments_label": "Number of Instruments",
        "cross_sectors": [
            (40, "Energy"), (41, "Industry"), (42, "Buildings"), (43, "Transport"),
            (44, "AFOLU"), (45, "Waste"), (46, "Cross-sectoral"),
        ],
        "cross_cats": [
            ("I", "Economic"), ("J", "Regulatory"),
            ("K", "Government investment and consumption"), ("L", "Information"),
            ("M", "Voluntary"),
        ],
    },
}


def write_summary_sheet(wb, lang: str) -> None:
    """Populate the Summary sheet with formulas referencing Approaches and Instruments overview."""
    config = LANG_CONFIG[lang]
    sheet_name = config.get("summary_sheet")
    if not sheet_name or sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    app = f"'{config['approaches_sheet']}'"
    ov = f"'{config['overview_sheet']}'"

    layout = SUMMARY_LAYOUT[lang]
    groups = layout["groups"]
    cats = layout["cats"]
    ov_cats = layout["ov_cats"]
    sectors = layout["sectors"]
    channels = layout["channels"]
    mitigations = layout["mitigations"]
    approaches_label = layout["approaches_label"]
    instruments_label = layout["instruments_label"]
    cross_sectors = layout["cross_sectors"]
    cross_cats = layout["cross_cats"]

    # ── Top stats (COUNTA on full columns minus header rows) ──
    ws["B4"] = f"=COUNTA({app}!C:C)-1"
    ws["B5"] = f"=COUNTA({ov}!A:A)-2"

    # ── Left side: Category/Group breakdown ──
    for row, group in groups:
        _set(ws, f"C{row}", f'=COUNTIF({app}!B:B, "{group}")')
        _set(ws, f"D{row}", f'=COUNTIF({ov}!J:J, "{group}")')

    # ── Right side: Category totals ──
    for row, cat in cats:
        _set(ws, f"I{row}", f'=COUNTIF({app}!A:A, "{cat}")')
    for row, cat in ov_cats:
        _set(ws, f"J{row}", f'=COUNTIF({ov}!I:I, "{cat}")')

    # ── Right side: Emission sectors ──
    sector_header = min(row for row, _ in sectors) - 1
    _set(ws, f"I{sector_header}", approaches_label)
    _set(ws, f"J{sector_header}", instruments_label)
    for row, sector in sectors:
        _set(ws, f"I{row}", f'=SUMPRODUCT((ISNUMBER(SEARCH("{sector}",{app}!F:F)))*1)')
        _set(ws, f"J{row}", f'=COUNTIF({ov}!K:K, "*{sector}*")')

    # ── Right side: Functioning channel ──
    channel_header = min(row for row, _ in channels) - 1
    _set(ws, f"J{channel_header}", instruments_label)
    for row, ch in channels:
        _set(ws, f"I{row}", f'=SUMPRODUCT((ISNUMBER(SEARCH("{ch}",{app}!H:H)))*1)')
        _set(ws, f"J{row}", f'=COUNTIF({ov}!M:M, "*{ch}*")')

    # ── Right side: Mitigation relevance ──
    mit_header = min(row for row, _ in mitigations) - 1
    _set(ws, f"I{mit_header}", approaches_label)
    _set(ws, f"J{mit_header}", instruments_label)
    for row, rel in mitigations:
        _set(ws, f"I{row}", f'=SUMPRODUCT((ISNUMBER(SEARCH("{rel}",{app}!G:G)))*1)')
        _set(ws, f"J{row}", f'=COUNTIF({ov}!H:H, "{rel}")')

    # ── Cross-tabulation: Emission Sector x Category (approaches) ──
    for row, sector in cross_sectors:
        for col_letter, cat_name in cross_cats:
            formula = f'=COUNTIFS({app}!F:F, "*{sector}*", {app}!A:A, "{cat_name}")'
            _set(ws, f"{col_letter}{row}", formula)
        _set(ws, f"N{row}", f"=SUM(I{row}:M{row})")

    # ── Uniform data-column widths within each table (mirrors the webpage module) ──
    # Left table: C/D. Right tables and cross-tab share I..N, so one group keeps
    # every table uniform; each group takes its widest current column.
    for group in (("C", "D"), ("I", "J", "K", "L", "M", "N")):
        defined = [
            ws.column_dimensions[col].width
            for col in group
            if col in ws.column_dimensions and ws.column_dimensions[col].width
        ]
        uniform = max(defined) if defined else 14.0
        for col in group:
            ws.column_dimensions[col].width = uniform

    # Position Summary as the first sheet
    if wb.sheetnames.index(sheet_name) != 0:
        wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index(sheet_name)))

    # ── Green data bars on all numeric cells ──
    # Solid (non-gradient) fill is applied in _apply_solid_databars via the
    # Office 2010 x14 extension, which openpyxl cannot write natively.
    blue_bar = DataBarRule(
        start_type="min", end_type="max",
        color=DATABAR_COLOR, showValue=True,
    )
    for ref in SUMMARY_DATABAR_RANGES[lang]:
        ws.conditional_formatting.add(ref, blue_bar)


def _set(ws, cell: str, value: str) -> None:
    """Write a value to a cell only if it's currently empty (preserves template static text)."""
    c = ws[cell]
    if c.value is None or (isinstance(c.value, str) and c.value.strip() == ""):
        c.value = value


def write_approaches_sheet(wb, lang: str, csv_data: list[tuple[list[str], list[dict[str, str]]]]) -> None:
    config = LANG_CONFIG[lang]
    sheet_name = config["approaches_sheet"]
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        if wb.sheetnames.index(sheet_name) != 1:
            wb._sheets.remove(ws)
            wb._sheets.insert(1, ws)
    else:
        ws = wb.create_sheet(sheet_name, 1)

    ws.append(config["approaches_headers"])
    lookup = load_chinese_approaches() if lang == "cn" else load_english_approaches()
    summaries: dict[str, dict[str, object]] = {}

    # Build CN→EN approach name mapping so Chinese-named instruments
    # merge with their English-named counterparts in the EN Approaches sheet.
    cn_to_en: dict[str, str] = {}
    if lang == "en":
        cn_approaches = load_chinese_approaches()
        seen_ids: set[int] = set()
        for cn_entry in cn_approaches.values():
            entry_id = id(cn_entry)
            if entry_id in seen_ids:
                continue
            seen_ids.add(entry_id)
            en_name = cn_entry.get("路径_en", "")
            cn_name = cn_entry.get("路径", "")
            if en_name and cn_name:
                cn_to_en[normalise_key(cn_name)] = en_name

    for csv_headers, rows in csv_data:
        for row in rows:
            approach = get_field(row, csv_headers, "Approach", "路径")
            if not approach:
                continue
            key = normalise_key(approach)
            # Normalize Chinese approach names to English for merging
            if lang == "en" and key in cn_to_en:
                approach = cn_to_en[key]
                key = normalise_key(approach)
            summary = summaries.setdefault(key, {"approach": approach, "sectors": [], "relevance": [], "channels": [], "ids": []})
            append_unique(summary["sectors"], get_field(row, csv_headers, "Emission sector", "排放部门"))
            append_unique(summary["relevance"], get_field(row, csv_headers, "Mitigation relevance", "减缓相关性"))
            append_unique(summary["channels"], get_field(row, csv_headers, "Functioning channel", "作用渠道"))
            append_unique(summary["ids"], get_field(row, csv_headers, "Policy Instrument ID", "政策工具ID"))

    def approach_sort_key(key: str) -> tuple[int, str]:
        classification = lookup.get(key, {})
        group = str(classification.get("Group", "") or classification.get("组别", ""))
        group_order = APPROACH_GROUP_SORT_ORDER.get(group.casefold(), 50)
        order = classification.get("_order")
        if order is not None:
            return (group_order, f"{int(order):04d}")
        return (group_order, str(summaries[key]["approach"]).casefold())

    for key in sorted(summaries, key=approach_sort_key):
        classification = lookup.get(key, {})
        summary = summaries[key]
        if lang == "cn":
            ws.append(
                [
                    classification.get("类别", ""),
                    classification.get("组别", ""),
                    classification.get("路径", str(summary["approach"])),
                    classification.get("缩写", approach_abbreviation(str(summary["approach"]))),
                    classification.get("定义", ""),
                    "；".join(summary["sectors"]),
                    "；".join(summary["relevance"]),
                    "；".join(summary["channels"]),
                    len(summary["ids"]),
                ]
            )
        else:
            ws.append(
                [
                    classification.get("Category", ""),
                    classification.get("Group", ""),
                    classification.get("Approach", str(summary["approach"])),
                    classification.get("Abbreviation", approach_abbreviation(str(summary["approach"]))),
                    classification.get("Definition", ""),
                    "; ".join(summary["sectors"]),
                    "; ".join(summary["relevance"]),
                    "; ".join(summary["channels"]),
                    len(summary["ids"]),
                ]
            )

    header_fill = PatternFill("solid", fgColor="FFE2EFDA")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, width in enumerate([18, 22, 36, 14, 100, 28, 20, 22, 24], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def _inject_x14_databars(sheet_xml: str, ranges: list[str]) -> str:
    """Rewrite data bar rules with the Office 2010 x14 extension for solid fill.

    openpyxl only writes the Excel 2007 base dataBar (always gradient-filled).
    Solid fill requires an x14:dataBar with gradient="0" plus a GUID link from
    the base cfRule, following the structure Excel/xlsxwriter produce.
    """
    range_guids = {}
    for index, ref in enumerate(ranges, start=1):
        range_guids[ref] = f"{{DA7ABA51-AAAA-BBBB-{index:04d}-000000000001}}"

    def link_rule(match: re.Match) -> str:
        sqref = match.group(1)
        guid = range_guids.get(sqref)
        if guid is None:
            return match.group(0)
        return (
            f'<conditionalFormatting sqref="{sqref}">'
            f'{match.group(2)}'
            f'<extLst><ext xmlns:x14="{X14_NS}" uri="{X14_CF_RULE_URI}">'
            f"<x14:id>{guid}</x14:id></ext></extLst>"
        )

    base_rule = re.compile(
        r'<conditionalFormatting sqref="([^"]+)">'
        r'(<cfRule type="dataBar"[^>]*><dataBar[^>]*>.*?</dataBar>)',
        re.DOTALL,
    )
    sheet_xml = base_rule.sub(link_rule, sheet_xml)

    blocks = []
    for ref in ranges:
        guid = range_guids[ref]
        blocks.append(
            f'<x14:conditionalFormatting xmlns:xm="{XM_NS}">'
            f'<x14:cfRule type="dataBar" id="{guid}">'
            f'<x14:dataBar minLength="0" maxLength="100" border="1" gradient="0" '
            f'negativeBarBorderColorSameAsPositive="0">'
            f'<x14:cfvo type="autoMin"/><x14:cfvo type="autoMax"/>'
            f'<x14:borderColor rgb="{DATABAR_COLOR}"/>'
            f'<x14:negativeFillColor rgb="FFFF0000"/>'
            f'<x14:negativeBorderColor rgb="FFFF0000"/>'
            f'<x14:axisColor rgb="FF000000"/>'
            f"</x14:dataBar></x14:cfRule>"
            f"<xm:sqref>{ref}</xm:sqref>"
            f"</x14:conditionalFormatting>"
        )
    ext_lst = (
        f'<extLst><ext xmlns:x14="{X14_NS}" uri="{X14_CFMTS_URI}">'
        f'<x14:conditionalFormattings>{"".join(blocks)}</x14:conditionalFormattings>'
        f"</ext></extLst>"
    )
    if "</worksheet>" not in sheet_xml:
        return sheet_xml + ext_lst
    return sheet_xml.replace("</worksheet>", ext_lst + "</worksheet>", 1)


def _apply_solid_databars(xlsx_path: Path, sheet_name: str, ranges: list[str]) -> None:
    """Post-process a saved xlsx to give the summary sheet solid-fill data bars."""
    import zipfile

    tmp_path = xlsx_path.with_name(xlsx_path.stem + "_tmp" + xlsx_path.suffix)
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rel_map = {}
        for entry in re.findall(r"<Relationship\b[^>]*>", rels_xml):
            rid_m = re.search(r'Id="([^"]+)"', entry)
            target_m = re.search(r'Target="([^"]+)"', entry)
            if rid_m and target_m:
                rel_map[rid_m.group(1)] = target_m.group(1)
        target = None
        for entry in re.findall(r"<sheet\b[^>]*>", wb_xml):
            name_m = re.search(r'name="([^"]+)"', entry)
            rid_m = re.search(r'r:id="([^"]+)"', entry)
            if name_m and rid_m and name_m.group(1) == sheet_name:
                rel_target = rel_map[rid_m.group(1)].lstrip("/")
                target = rel_target if rel_target.startswith("xl/") else "xl/" + rel_target
                break
        if target is None:
            tmp_path.unlink()
            return

        sheet_xml = zin.read(target).decode("utf-8")
        if X14_CFMTS_URI not in sheet_xml:
            sheet_xml = _inject_x14_databars(sheet_xml, ranges)

        for item in zin.infolist():
            data = sheet_xml.encode("utf-8") if item.filename == target else zin.read(item.filename)
            zout.writestr(item, data)
    tmp_path.replace(xlsx_path)


def normalise_workbook_fonts(wb: Workbook, font_name: str) -> None:
    """Force one font family across every sheet; script-written rows default to Calibri.

    Fonts are reassigned unconditionally (no name check) because theme-dependent
    fonts (scheme="minor") keep their name but render with the theme's CJK typeface,
    and skipped cells otherwise save with stale style ids.
    """
    from openpyxl.cell.cell import MergedCell

    for ws in wb.worksheets:
        # Merged-cell members can't take a font directly; unmerge, normalise, re-merge
        merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
        for rng in merged_ranges:
            ws.unmerge_cells(rng)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                font = cell.font
                cell.font = Font(
                    name=font_name,
                    sz=font.sz,
                    bold=font.bold,
                    italic=font.italic,
                    vertAlign=font.vertAlign,
                    underline=font.underline,
                    strike=font.strike,
                    color=font.color,
                )
        for rng in merged_ranges:
            ws.merge_cells(rng)
        # Drop merged-member records that carry no visible style (borders or fills).
        # Members with borders must survive: Excel draws merged-region edges from the
        # member cells, and openpyxl only writes cells present in _cells.
        for coord, cell in list(ws._cells.items()):
            if isinstance(cell, MergedCell):
                border = cell.border
                has_border = any(
                    side is not None and side.style
                    for side in (border.left, border.right, border.top, border.bottom, border.diagonal)
                )
                has_fill = cell.fill is not None and bool(cell.fill.patternType)
                if not has_border and not has_fill:
                    del ws._cells[coord]

    # Default cell style font (fontId 0) — governs cells without an explicit style
    if wb._fonts and wb._fonts[0] is not None:
        default = wb._fonts[0]
        wb._fonts[0] = Font(name=font_name, sz=default.sz or 11)

    # Theme minor font: CJK text in theme-dependent cells falls back to the
    # Hans script typeface, so fix it for the Chinese workbook
    theme = getattr(wb, "loaded_theme", None)
    if theme and font_name == "等线":
        if isinstance(theme, bytes):
            theme = theme.decode("utf-8")
        theme = theme.replace('<a:latin typeface="Calibri"/>', f'<a:latin typeface="{font_name}"/>')
        theme = theme.replace('<a:ea typeface=""/>', f'<a:ea typeface="{font_name}"/>')
        theme = theme.replace('script="Hans" typeface="宋体"', f'script="Hans" typeface="{font_name}"')
        wb.loaded_theme = theme.encode("utf-8")


def apply_sheet_order(wb: Workbook, lang: str) -> None:
    """Set the final sheet order: front matter, Summary/Approaches/overview,
    data sheets, any unlisted template sheets, then back matter."""
    config = LANG_CONFIG[lang]
    front = config.get("front_sheets", [])
    back = config.get("back_sheets", [])
    special = [
        sheet
        for sheet in (
            config.get("summary_sheet"),
            config.get("approaches_sheet"),
            config.get("overview_sheet"),
        )
        if sheet
    ]
    data_sheets = list(config["sheets"].values())
    desired = set(front + special + data_sheets + back)
    extras = [ws.title for ws in wb.worksheets if ws.title not in desired]
    order = front + special + data_sheets + extras + back
    wb._sheets = [wb[title] for title in order if title in wb.sheetnames]


ENDED_SHEET_COLUMNS = {
    "en": {
        "headers": [
            "English name",
            "Domestic instrument name",
            "Status",
            "Start year",
            "End year",
            "Instrument category",
            "Instrument group",
            "Emission sector",
            "Sub-sector",
            "Legal document",
            "Weblinks",
        ],
        "fields": [
            "English instrument name",
            "Domestic instrument name",
            "Status",
            "Start date",
            "End date",
            "Group",
            "Emission sector",
            "Sub-sector",
            "Legal document",
            "Other weblinks",
        ],
        "categories": {
            "economic_instruments": "Economic",
            "regulatory_instruments": "Regulatory",
            "government_i_c": "Government investment and consumption",
            "information_instruments": "Information",
            "voluntary_approaches": "Voluntary",
        },
    },
    "cn": {
        "headers": [
            "英文名",
            "本国工具名称",
            "状态",
            "开始年份",
            "结束年份",
            "工具类别",
            "工具组别",
            "排放部门",
            "子行业",
            "法律文件名称",
            "网页链接",
        ],
        "fields": [
            "英文工具名称",
            "本国工具名称",
            "状态",
            "生效日期",
            "终止日期",
            "组别",
            "排放部门",
            "子行业",
            "法律文件名称",
            "其他网页链接",
        ],
        "categories": {
            "economic_instruments": "经济工具",
            "regulatory_instruments": "规制工具",
            "government_i_c": "政府投资与消费",
            "information_instruments": "信息工具",
            "voluntary_approaches": "自愿措施",
        },
    },
}


def year_from_date(value: str) -> str:
    text = str(value or "").strip()
    if not text or text.upper() == "N/A":
        return text
    if "/" in text:
        return text.split("/")[-1]
    return text[:4]


def write_ended_sheet(wb: Workbook, lang: str, outputs_dir: Path) -> None:
    """Append instruments moved to *_ended.csv files below the curated rows
    kept in the template's ended sheet."""
    config = LANG_CONFIG[lang]
    sheet_name = config.get("ended_sheet")
    if not sheet_name:
        return
    column_config = ENDED_SHEET_COLUMNS[lang]
    fields = column_config["fields"]
    categories = column_config["categories"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(column_config["headers"])

    # Preserve the curated rows kept in the template; clear everything below them
    last_curated = 1
    for row_idx in range(2, ws.max_row + 1):
        if any(ws.cell(row=row_idx, column=col).value not in (None, "") for col in range(1, 12)):
            last_curated = row_idx
    if ws.max_row > last_curated:
        ws.delete_rows(last_curated + 1, ws.max_row - last_curated)

    curated_names = {
        str(ws.cell(row=row_idx, column=1).value or "").strip().casefold()
        for row_idx in range(2, last_curated + 1)
    }

    generated: list[list[str]] = []
    for csv_path in sorted(outputs_dir.glob(f"CCPID_{lang}_*_ended.csv")):
        key = csv_path.stem.removeprefix(f"CCPID_{lang}_").removesuffix("_ended")
        category = categories.get(key)
        if category is None:
            continue
        csv_headers, csv_rows = read_csv(csv_path)
        for row in csv_rows:
            name = get_field(row, csv_headers, fields[0])
            if not name or name.strip().casefold() in curated_names:
                continue
            status = get_field(row, csv_headers, fields[2])
            if lang == "en" and status.strip().casefold() == "ended":
                status = "Ended"
            generated.append(
                [
                    name,
                    get_field(row, csv_headers, fields[1]),
                    status,
                    year_from_date(get_field(row, csv_headers, fields[3])),
                    year_from_date(get_field(row, csv_headers, fields[4])),
                    category,
                    get_field(row, csv_headers, fields[5]),
                    get_field(row, csv_headers, fields[6]),
                    get_field(row, csv_headers, fields[7]),
                    get_field(row, csv_headers, fields[8]),
                    get_field(row, csv_headers, fields[9]),
                ]
            )

    for offset, values in enumerate(generated, start=1):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=last_curated + offset, column=col_idx, value=value)

    # Header row bold; all data rows non-bold
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.font.bold:
                cell.font = cell.font.copy(bold=False)


def update_cover_date(wb: Workbook, lang: str) -> None:
    """Refresh the cover sheet's 'as of' date to the current month on each export."""
    cover_sheet = LANG_CONFIG[lang].get("cover_sheet")
    if not cover_sheet or cover_sheet not in wb.sheetnames:
        return
    today = datetime.date.today()
    if lang == "cn":
        prefix, text = "截至", f"截至{today.year}年{today.month}月"
    else:
        prefix, text = "as of ", f"as of {today.strftime('%B')} {today.year}"
    for row in wb[cover_sheet].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().startswith(prefix):
                cell.value = text
                return


def export_dataset(lang: str, outputs_dir: Path) -> Path | None:
    config = LANG_CONFIG[lang]
    template_path = config["template"]
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    csv_files = sorted(outputs_dir.glob(f"CCPID_{lang}_*.csv"))
    if not csv_files:
        return None

    wb = load_workbook(template_path)
    loaded_csv_data: list[tuple[list[str], list[dict[str, str]]]] = []
    for csv_path in csv_files:
        key = csv_path.stem.removeprefix(f"CCPID_{lang}_")
        sheet_name = config["sheets"].get(key)
        if not sheet_name:
            continue
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{sheet_name}' not found in {template_path}")
        headers, rows = read_csv(csv_path)
        loaded_csv_data.append((headers, rows))
        write_rows(wb[sheet_name], headers, rows)

    write_summary_sheet(wb, lang)
    write_approaches_sheet(wb, lang, loaded_csv_data)
    write_instruments_overview_sheet(wb, lang)
    write_ended_sheet(wb, lang, outputs_dir)
    update_cover_date(wb, lang)

    apply_sheet_order(wb, lang)
    normalise_workbook_fonts(wb, "等线" if lang == "cn" else "Calibri")

    output_path = outputs_dir / str(config["output"])
    wb.save(output_path)
    summary_sheet = config.get("summary_sheet")
    if summary_sheet:
        _apply_solid_databars(output_path, summary_sheet, SUMMARY_DATABAR_RANGES[lang])
    return output_path


def export_evidence_log(outputs_dir: Path) -> Path | None:
    csv_path = outputs_dir / "evidence_log.csv"
    if not csv_path.exists():
        return None

    headers, rows = read_csv(csv_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "evidence_log"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    output_path = outputs_dir / "evidence_log.xlsx"
    wb.save(output_path)
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--outputs-dir", type=Path, default=DEFAULT_OUTPUTS_DIR)
    parser.add_argument("--lang", choices=["cn", "en", "all"], default="all")
    parser.add_argument("--skip-evidence", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    langs = ["cn", "en"] if args.lang == "all" else [args.lang]
    exported: list[Path] = []
    for lang in langs:
        path = export_dataset(lang, args.outputs_dir)
        if path:
            exported.append(path)
    if not args.skip_evidence:
        evidence = export_evidence_log(args.outputs_dir)
        if evidence:
            exported.append(evidence)

    if exported:
        for path in exported:
            print(f"Exported {path}")
    else:
        print("No CSV outputs found to export.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
