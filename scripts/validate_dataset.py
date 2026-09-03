#!/usr/bin/env python3
"""Validate IFCMA dataset exports in the outputs directory.

The validator is intentionally dependency-light. It reads CSV, TSV, JSON, JSONL,
and Markdown pipe tables with the Python standard library. XLSX support is used
only when openpyxl is installed.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import ssl
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from policy_id import expected_id_codes, parse_policy_id


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUTS_DIR = ROOT / "outputs"
DEFAULT_REPORT_PATH = ROOT / "logs" / "validation_report.md"
APPROACHES_PATH = ROOT / "inputs" / "classification" / "approaches_en.md"
APPROACHES_CN_PATH = ROOT / "inputs" / "classification" / "approaches_cn.md"
CN_TEMPLATE_PATH = ROOT / "inputs" / "template_cn.xlsx"

REQUIRED_FIELD_GROUPS = [
    ("Policy Instrument ID", ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"]),
    ("Instrument / subscheme", ["Instrument / subscheme", "Instrument / Subscheme", "row_type", "instrument_subscheme"]),
    ("Group", ["Group", "IFCMA type", "Instrument Type", "instrument_type", "type"]),
    ("Approach", ["Approach", "IFCMA approach", "approach"]),
    ("Country", ["Country", "country"]),
    ("Jurisdiction level", ["Jurisdiction level", "jurisdiction_level"]),
    ("Status", ["Status", "status"]),
]

NAME_FIELD_GROUPS = [
    ["Domestic instrument name", "Domestic name", "domestic_instrument_name", "domestic_name"],
    ["English instrument name", "English name", "english_instrument_name", "english_name"],
]

CATEGORY_FIELDS = ["Category", "IFCMA category", "Instrument Category", "category", "instrument_category"]
GROUP_FIELDS = ["Group", "IFCMA type", "Instrument Type", "instrument_type", "type"]
APPROACH_FIELDS = ["Approach", "IFCMA approach", "approach"]
STATUS_FIELDS = ["Status", "status"]
START_DATE_FIELDS = ["Start date", "start_date"]
END_DATE_FIELDS = ["End date", "end_date"]
SOURCE_URL_FIELDS = ["source_url", "Source URL", "Legal document", "legal_document", "Other weblinks", "Other relevant websites"]
LEGAL_DOCUMENT_FIELDS = ["Legal document", "legal_document"]
CONFIDENCE_FIELDS = ["confidence_score", "Confidence score", "confidence", "Confidence"]
REVIEW_FIELDS = ["needs_human_review", "Needs human review", "human_review", "Human review"]
# Instrument-specific and GHG fields expected to contain quantitative data.
# If these fields have numeric content, they must include an inline source
# attribution (CN: 来源, EN: Source:).
INSTRUMENT_QUANTITATIVE_FIELDS = [
    "Tax and Tax Incentive: annual revenue",
    "Tax and Tax Incentive: annual revenue forgone",
    "Subsidy: annual budget/ expenditure",
    "Subsidy: Limit",
    "Trading System: cap",
    "Trading System: Free Allowance",
    "Trading System: revenue (annual)",
    "Trading System: Volume",
    "GHG emission coverage (absolute)",
    "GHG emission coverage (% domestic emissions)",
]

CONCRETE_DATA_PATTERN = re.compile(r"\d")

SOURCE_MARKER_CN = "来源"
SOURCE_MARKER_EN = "source:"
INTENSITY_VALUE_FIELDS = ["Intensity (Value)"]
INTENSITY_DETAIL_FIELDS = ["Intensity (Details)"]
COMPLIANCE_I_FIELDS = ["Compliance calculation methodology I"]
COMPLIANCE_II_FIELDS = ["Compliance calculation methodology II"]
PARENT_FIELDS = [
    "parent_instrument_id",
    "Parent instrument ID",
    "Parent Policy Instrument ID",
    "parent_policy_instrument_id",
    "parent_id",
]
ROW_TYPE_FIELDS = ["Instrument / subscheme", "Instrument / Subscheme", "row_type", "instrument_subscheme"]
RENEWABLE_REC_APPROACH = "tradable renewable electricity credits, quota or tradable performance standards"

VALID_STATUSES = {"in force", "scheduled", "ended", "non-existent", "生效", "计划实施", "已终止", "不存在"}

VALUE_ALIASES = {
    "工具": "Instrument",
    "子方案": "Subscheme",
    "交易机制": "Trading scheme",
    "补贴": "Subsidy",
    "税收": "Tax",
    "排放交易体系（ets）": "Emissions trading system",
    "排放交易体系(ets)": "Emissions trading system",
    "可交易绩效标准": "Tradable performance standards",
    "可交易可再生能源证书/配额/rps": "Tradable renewable electricity credits, quota or tradable performance standards",
    "优惠贷款、贷款担保和信贷支持": "Concessional loans, loan guarantees and credit support",
    "可再生电力差价合约": "Renewable electricity contract for difference",
    "可再生电力差价合约（cfd）": "Renewable electricity contract for difference",
    "可再生电力差价合约(cfd)": "Renewable electricity contract for difference",
    "低碳车辆购置补贴": "Vehicle purchase subsidy",
    "车辆购置税优惠": "Vehicle purchase tax incentive",
    "生态系统服务付费": "Ecosystem service payments",
    "以旧换新补贴": "Trade-in subsidy",
    "清洁取暖补贴": "Clean heating subsidy",
    "回收处理奖补": "Recycling and treatment subsidy",
    "低排放机械购置补贴": "Low-emission machinery purchase subsidy",
    "中央预算内投资": "Central budget investment",
    "政府投资基金": "Government investment fund",
    "政府资助研发与示范计划": "Government-funded RD&D programmes",
    "公共投资": "Public investment",
    "公共采购": "Public procurement",
    "绿色公共采购": "Green public procurement",
    "中国": "CHN",
    "国家": "national",
    "供给侧": "Supply-side",
    "需求侧": "demand-side",
    "环境": "environment",
    "生效": "in force",
    "计划实施": "scheduled",
    "燃料消费税": "Fuel excise tax",
    "差别化车辆税": "Differentiated vehicle tax",
    "车船税优惠": "Vehicle ownership tax incentive",
    "增值税优惠": "VAT incentive",
    "环境污染税": "Environmental pollution tax",
    "企业所得税优惠": "CIT incentive",
    "消费税优惠": "Consumption tax incentive",
    "政府性基金减免": "Parafiscal levy exemption",
    "城镇土地使用税优惠": "Urban land use tax incentive",
    "已终止": "ended",
}

INSTRUMENT_SPECIFIC_FIELD_PREFIXES = {
    "Tax": (
        "Subsidy:",
        "Trading System:",
        "补贴：",
        "交易系统：",
    ),
    "Subsidy": (
        "Tax and Tax Incentive:",
        "Trading System:",
        "税收及激励：",
        "交易系统：",
    ),
    "Trading scheme": (
        "Tax and Tax Incentive:",
        "Subsidy:",
        "税收及激励：",
        "补贴：",
    ),
    "Public investment": (
        "Public procurement:",
        "公共采购：",
    ),
    "Public procurement": (
        "Public investment: Co-financing",
        "Public investment: Selection process",
        "公共投资：共同融资",
        "公共投资：选择流程",
    ),
    "Public appraisal rules": (
        "Public investment:",
        "Public procurement:",
        "公共投资：",
        "公共采购：",
    ),
}

# Conservative official-domain defaults. Override/extend with
# --official-domain-suffix for jurisdiction-specific validation.
DEFAULT_OFFICIAL_DOMAIN_SUFFIXES = (
    ".gov",
    ".gov.cn",
    ".gob",
    ".gouv",
    ".go.jp",
    ".go.kr",
    ".gov.uk",
    ".europa.eu",
    ".unfccc.int",
    ".oecd.org",
)

# Official-looking portals that are useful for leads but should not be treated
# as the canonical final policy source when a formal ministry notice page exists.
CONSULTATION_OR_LEAD_HOSTS = {
    "yyglxxbsgw.ndrc.gov.cn",
}

CJK_PATTERN = re.compile(r"[⺀-⻿　-〿㐀-䶿一-鿿豈-﫿＀-￯]")

# Schema Activity allowed_values — multi-term options where "or" means pick
# applicable terms, not write all. Tuples of (canonical, [canonical_terms]).
_ACTIVITY_SCHEMA_OPTIONS: list[tuple[str, list[str]]] = [
    ("Conservation, protection or storage", ["conservation", "protection", "storage"]),
    ("Extraction, abstraction or harvesting", ["extraction", "abstraction", "harvesting"]),
    ("Production, generation or conversion", ["production", "generation", "conversion"]),
    ("Abatement or prevention", ["abatement", "prevention"]),
    ("Export, import, sales or purchase", ["export", "import", "sales", "purchase"]),
    ("Registration, licensing or other administrative tasks", ["registration", "licensing"]),
    ("Consumption (use, ownership or capital formation)", ["ownership", "capital formation"]),
    ("Disposal, collection or sorting after use", ["disposal", "collection", "sorting"]),
    ("Recycling, repurposing or other treatment after use", ["recycling", "repurposing", "treatment"]),
]

# Synonyms for cross-segment duplication detection — same concept, different word.
_ACTIVITY_SYNONYM_GROUPS: list[set[str]] = [
    {"recycling", "recovery"},
    {"repurposing", "reuse"},
    {"purchase", "sales"},
    {"conservation", "protection"},
]


def _activity_vocab_terms(segment: str) -> set[str]:
    """Extract significant lowercase terms from an Activity segment, with synonym canonicalization."""
    cleaned = re.sub(r"\([^)]*\)", "", segment.lower()).strip()
    cleaned = re.sub(r"[;,/]", " ", cleaned)
    stop = {"or", "and", "other", "after", "use", "post", "consumer",
            "post-consumer", "post-consumption", "ownership"}
    terms = {t for t in cleaned.split() if t not in stop}
    # Normalize synonyms to canonical forms
    for group in _ACTIVITY_SYNONYM_GROUPS:
        if terms & group:
            canonical = sorted(group)[0]
            terms = (terms - group) | {canonical}
    return terms


def _canonical_option_terms(terms: list[str]) -> set[str]:
    """Apply the same synonym normalization to schema option terms."""
    result = set(terms)
    for group in _ACTIVITY_SYNONYM_GROUPS:
        if result & group:
            canonical = sorted(group)[0]
            result = (result - group) | {canonical}
    return result


def _check_activity_vocabulary(activity_value: str) -> list[str]:
    """Warn when Activity uses all terms from a schema option or duplicates terms across segments."""
    if not activity_value or activity_value.strip().upper() == "N/A":
        return []
    segments = [s.strip() for s in re.split(r";|；", activity_value) if s.strip()]
    if not segments:
        return []
    warnings: list[str] = []

    # Check for duplicate schema-vocabulary terms across segments.
    # Only flag terms that belong to known schema options — generic words
    # like "discharge" in "discharge of air/water/solid" are legitimate.
    if len(segments) >= 2:
        all_schema_terms: set[str] = set()
        for _name, terms in _ACTIVITY_SCHEMA_OPTIONS:
            all_schema_terms.update(_canonical_option_terms(terms))
        term_seen: dict[str, int] = {}
        for i, segment in enumerate(segments):
            for t in _activity_vocab_terms(segment):
                if t not in all_schema_terms:
                    continue
                if t not in term_seen:
                    term_seen[t] = i
                elif term_seen[t] != i:
                    warnings.append(
                        f"\"{t}\" appears in multiple Activity segments — "
                        f"check whether it belongs in only one segment"
                    )
                    term_seen[t] = -1

    return warnings


ISIC_DIVISION_RANGES = {
    "A": range(1, 4),
    "B": range(5, 10),
    "C": range(10, 34),
    "D": range(35, 36),
    "E": range(36, 40),
    "F": range(41, 44),
    "G": range(45, 48),
    "H": range(49, 54),
    "I": range(55, 57),
    "J": range(58, 64),
    "K": range(64, 67),
    "L": range(68, 69),
    "M": range(69, 76),
    "N": range(77, 83),
    "O": range(84, 85),
    "P": range(85, 86),
    "Q": range(86, 89),
    "R": range(90, 94),
    "S": range(94, 97),
    "T": range(97, 99),
    "U": range(99, 100),
}

CN_SHEET_BY_TEMPLATE = {
    "Economic instruments": "经济工具",
    "Regulatory instruments": "规制工具",
    "Government I&C": "政府投资与消费",
    "Information instruments": "信息工具",
    "Voluntary approaches": "自愿措施",
}


def normalise_header(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def cell_ref_column(ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", ref.upper())
    value = 0
    for letter in letters:
        value = value * 26 + (ord(letter) - ord("A") + 1)
    return value


def load_xlsx_row_headers(path: Path, row_number: int = 2) -> dict[str, list[str]]:
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    if not path.exists():
        return {}
    with zipfile.ZipFile(path) as archive:
        shared_strings: list[str] = []
        try:
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall("a:si", ns):
                shared_strings.append("".join(t.text or "" for t in item.findall(".//a:t", ns)))
        except KeyError:
            pass
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        headers_by_sheet: dict[str, list[str]] = {}
        sheets = workbook.find("a:sheets", ns)
        if sheets is None:
            return headers_by_sheet
        for sheet in sheets:
            sheet_name = sheet.attrib["name"]
            rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = relmap[rel_id]
            if target.startswith("/xl/"):
                worksheet = ET.fromstring(archive.read(target[1:]))
            else:
                worksheet = ET.fromstring(archive.read("xl/" + target))
            rows = worksheet.findall(".//a:sheetData/a:row", ns)
            selected = next((row for row in rows if row.attrib.get("r") == str(row_number)), None)
            if selected is None:
                continue
            values_by_index: dict[int, str] = {}
            for cell in selected.findall("a:c", ns):
                cell_type = cell.attrib.get("t")
                if cell_type == "s":
                    raw_value = cell.find("a:v", ns)
                    value = raw_value.text if raw_value is not None else ""
                    if value:
                        value = shared_strings[int(value)]
                elif cell_type == "inlineStr":
                    is_el = cell.find("a:is", ns)
                    value = "".join(t.text or "" for t in is_el.findall(".//a:t", ns)) if is_el is not None else ""
                else:
                    raw_value = cell.find("a:v", ns)
                    value = raw_value.text if raw_value is not None else ""
                values_by_index[cell_ref_column(cell.attrib.get("r", ""))] = value
            max_index = max(values_by_index, default=0)
            headers_by_sheet[sheet_name] = [values_by_index.get(index, "") for index in range(1, max_index + 1)]
        return headers_by_sheet


def load_cn_field_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    schema_path = ROOT / "rules" / "schema.yaml"
    if not schema_path.exists():
        return aliases
    cn_headers = load_xlsx_row_headers(CN_TEMPLATE_PATH)
    schema_text = schema_path.read_text(encoding="utf-8-sig")
    for template, sheet in CN_SHEET_BY_TEMPLATE.items():
        match = re.search(rf"  {re.escape(template)}:\n(?:    .+\n)*?    columns:\n((?:      - .+\n)+)", schema_text)
        if not match:
            continue
        english = [line.split("- ", 1)[1].strip().strip('"') for line in match.group(1).splitlines()]
        for source, target in zip(english, cn_headers.get(sheet, [])):
            aliases[normalise_header(source)] = target
    return aliases


CN_FIELD_ALIASES = load_cn_field_aliases()


def normalise_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"", "n/a", "na", "not applicable", "not specified", "none", "null"}:
        return ""
    return text


def truthy(value: object) -> bool:
    return normalise_value(value).strip().lower() in {"true", "yes", "y", "1", "review", "needs review"}


def find_field(row: dict[str, object], names: list[str]) -> str | None:
    by_norm = {normalise_header(k): k for k in row.keys()}
    for name in names:
        key = by_norm.get(normalise_header(name))
        if key is None:
            alias = CN_FIELD_ALIASES.get(normalise_header(name))
            key = by_norm.get(normalise_header(alias or ""))
        if key is not None:
            return key
    return None


def get_value(row: dict[str, object], names: list[str]) -> str:
    key = find_field(row, names)
    value = normalise_value(row.get(key)) if key is not None else ""
    return VALUE_ALIASES.get(value.strip().lower(), value)


def get_raw_value(row: dict[str, object], names: list[str]) -> str:
    """Like get_value but without CN→EN value translation."""
    key = find_field(row, names)
    return normalise_value(row.get(key)) if key is not None else ""


def split_multi(value: str) -> list[str]:
    return [part.strip() for part in re.split(r";|；|\|", value) if part.strip()]


def is_valid_isic_division(value: str) -> bool:
    match = re.fullmatch(r"([A-U])(\d{2})", value.strip().upper())
    if not match:
        return False
    section, division = match.group(1), int(match.group(2))
    return division in ISIC_DIVISION_RANGES.get(section, range(0))


def is_blank_allowed(field: str, group: str) -> bool:
    return any(field.startswith(prefix) for prefix in INSTRUMENT_SPECIFIC_FIELD_PREFIXES.get(group, ()))


def read_markdown_table(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    table_lines = [line.strip() for line in lines if line.strip().startswith("|") and line.strip().endswith("|")]
    if len(table_lines) < 2:
        return []
    header = [cell.strip() for cell in table_lines[0].strip("|").split("|")]
    rows = []
    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if len(cells) != len(header):
            continue
        rows.append(dict(zip(header, cells)))
    return rows


def read_dataset(path: Path) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if suffix == ".tsv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle, delimiter="\t"))
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(data, list):
            return [row for row in data if isinstance(row, dict)]
        if isinstance(data, dict):
            for key in ("rows", "data", "records", "instruments"):
                if isinstance(data.get(key), list):
                    return [row for row in data[key] if isinstance(row, dict)]
        return []
    if suffix == ".jsonl":
        rows = []
        for line in path.read_text(encoding="utf-8-sig").splitlines():
            if line.strip():
                row = json.loads(line)
                if isinstance(row, dict):
                    rows.append(row)
        return rows
    if suffix in {".md", ".markdown"}:
        return read_markdown_table(path)
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            raise RuntimeError("XLSX validation requires openpyxl; install it or export to CSV.")
        wb = load_workbook(path, read_only=True, data_only=True)
        rows: list[dict[str, object]] = []
        for ws in wb.worksheets:
            values = list(ws.iter_rows(values_only=True))
            if not values:
                continue
            first = [str(cell).strip() if cell is not None else "" for cell in values[0]]
            second = [str(cell).strip() if cell is not None else "" for cell in values[1]] if len(values) > 1 else []
            if (
                len(values) > 1
                and not get_value(dict.fromkeys(first, ""), ["Policy Instrument ID"])
                and find_field(dict.fromkeys(second, ""), ["Policy Instrument ID"]) is not None
            ):
                headers = second
                data_values = values[2:]
            else:
                headers = first
                data_values = values[1:]
            for raw in data_values:
                if not any(cell is not None and str(cell).strip() for cell in raw):
                    continue
                rows.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]})
        return rows
    return []


def load_allowed_classification() -> tuple[set[str], dict[str, set[str]], set[str]]:
    categories: set[str] = set()
    groups_by_category: dict[str, set[str]] = defaultdict(set)
    approaches: set[str] = set()
    if not APPROACHES_PATH.exists():
        return categories, groups_by_category, approaches
    for row in read_markdown_table(APPROACHES_PATH):
        category = normalise_value(row.get("Category"))
        group = normalise_value(row.get("Group"))
        approach = normalise_value(row.get("Approach"))
        if category:
            categories.add(category)
        if category and group:
            groups_by_category[category].add(group)
        if approach:
            approaches.add(approach)
    return categories, groups_by_category, approaches


def load_cn_classification() -> tuple[set[str], set[str]]:
    """Parse approaches_cn.md: CN group names from ### headers, CN approach names from table rows."""
    cn_groups: set[str] = set()
    cn_approaches: set[str] = set()
    if not APPROACHES_CN_PATH.exists():
        return cn_groups, cn_approaches
    for line in APPROACHES_CN_PATH.read_text(encoding="utf-8").splitlines():
        if line.startswith("### "):
            parts = line[4:].strip().rsplit(maxsplit=1)
            if len(parts) == 2:
                cn_groups.add(parts[1].strip())
        elif line.startswith("|") and not line.startswith("|---"):
            cells = [cell.strip() for cell in line.strip("|").split("|")]
            if len(cells) >= 2 and cells[0] and cells[1] and cells[0] != "Approach":
                cn_approaches.add(cells[1])
    return cn_groups, cn_approaches


def parse_date(value: str) -> dt.date | None:
    value = normalise_value(value)
    if not value:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%Y", "%Y"):
        try:
            parsed = dt.datetime.strptime(value, fmt)
            if fmt == "%Y":
                return dt.date(parsed.year, 1, 1)
            if fmt == "%m/%Y":
                return dt.date(parsed.year, parsed.month, 1)
            return parsed.date()
        except ValueError:
            continue
    return None


def official_url(url: str, suffixes: tuple[str, ...]) -> bool:
    parsed = urlparse(url.strip())
    host = parsed.netloc.lower()
    if not parsed.scheme or parsed.scheme not in {"http", "https"} or not host:
        return False
    host = host.split("@")[-1].split(":")[0]
    return any(host == suffix.lstrip(".") or host.endswith(suffix) for suffix in suffixes)


def url_host(url: str) -> str:
    parsed = urlparse(url.strip())
    return parsed.netloc.lower().split("@")[-1].split(":")[0]


def consultation_or_lead_url(url: str) -> bool:
    return url_host(url) in CONSULTATION_OR_LEAD_HOSTS


def check_url_health(url: str, timeout: int = 10) -> tuple[int, str | None]:
    """Check URL reachability. Returns (status_code, error_message).

    Tries HEAD first; falls back to GET (Range: bytes=0-0) for 405 responses.
    """
    ctx = ssl.create_default_context()
    headers = {"User-Agent": "CCPID-validator/1.0"}
    for method in ("HEAD", "GET"):
        try:
            req = Request(url, method=method, headers=headers)
            if method == "GET":
                req.add_header("Range", "bytes=0-0")
            with urlopen(req, timeout=timeout, context=ctx) as resp:
                return resp.status, None
        except Exception as exc:
            if method == "HEAD":
                continue  # try GET
            return 0, str(exc)
    return 0, "HEAD and GET both failed"


def iter_urls(value: str) -> list[str]:
    return re.findall(r"https?://[^\s,;|)>\]]+", value)


def infer_parent_id(instrument_id: str) -> str:
    if re.search(r"S\d+$", instrument_id):
        return re.sub(r"S\d+$", "S000", instrument_id)
    return ""


def normalise_control_value(value: str) -> str:
    stripped = value.strip()
    return VALUE_ALIASES.get(stripped.casefold(), stripped)


def validate_rows(
    rows_by_file: dict[Path, list[dict[str, object]]],
    official_suffixes: tuple[str, ...],
    check_urls: bool = False,
) -> tuple[list[dict[str, str]], Counter[str]]:
    valid_categories, valid_groups_by_category, valid_approaches = load_allowed_classification()
    valid_cn_groups, valid_cn_approaches = load_cn_classification()
    issues: list[dict[str, str]] = []
    stats: Counter[str] = Counter()
    all_ids: set[str] = set()
    all_ids_by_file: dict[Path, set[str]] = defaultdict(set)
    id_locations: dict[tuple[Path, str], list[str]] = defaultdict(list)
    duplicate_keys: dict[tuple[Path, str, str, str], list[str]] = defaultdict(list)
    url_health_cache: dict[str, tuple[int, str | None]] = {}

    flattened: list[tuple[Path, int, dict[str, object]]] = []
    for path, rows in rows_by_file.items():
        for index, row in enumerate(rows, start=2):
            flattened.append((path, index, row))
            instrument_id = get_value(row, ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"])
            if instrument_id:
                all_ids.add(instrument_id)
                all_ids_by_file[path].add(instrument_id)
                id_locations[(path, instrument_id)].append(f"{path.name}:{index}")
            country = get_value(row, ["Country", "country"]).casefold()
            domestic = get_value(row, NAME_FIELD_GROUPS[0]).casefold()
            english = get_value(row, NAME_FIELD_GROUPS[1]).casefold()
            name = domestic or english
            if country and name:
                duplicate_keys[(path, country, name, get_value(row, GROUP_FIELDS).casefold())].append(f"{path.name}:{index}")

    def add(path: Path, row_number: int, check: str, detail: str, severity: str = "error") -> None:
        issues.append(
            {
                "severity": severity,
                "file": path.name,
                "row": str(row_number),
                "check": check,
                "detail": detail,
            }
        )
        stats[check] += 1

    for path, row_number, row in flattened:
        stats["rows_checked"] += 1

        group = get_value(row, GROUP_FIELDS)
        blank_fields = [
            field
            for field, value in row.items()
            if field and (value is None or not str(value).strip()) and not is_blank_allowed(field, group)
        ]
        if blank_fields:
            add(path, row_number, "blank template cells", f"Blank fields: {', '.join(blank_fields[:20])}")

        if path.name.startswith("CCPID_cn"):
            english_not_found = [
                field for field, value in row.items() if str(value).strip().lower() == "not found"
            ]
            if english_not_found:
                add(path, row_number, "untranslated not found", f"Use 未找到 in Chinese output fields: {', '.join(english_not_found[:20])}")

            missing_source_labels = []
            for field_name in INSTRUMENT_QUANTITATIVE_FIELDS:
                key = find_field(row, [field_name])
                value = normalise_value(row.get(key)) if key else ""
                if value and CONCRETE_DATA_PATTERN.search(value) and SOURCE_MARKER_CN not in value:
                    missing_source_labels.append(key or field_name)
            if missing_source_labels:
                add(
                    path,
                    row_number,
                    "instrument quantitative source missing",
                    f"Instrument-specific/GHG fields with numeric data missing source attribution (add 来源: ...): {', '.join(missing_source_labels[:20])}",
                    "warning",
                )

        if path.name.startswith("CCPID_en"):
            for field in row:
                value = str(row.get(field, ""))
                if CJK_PATTERN.search(value):
                    add(
                        path,
                        row_number,
                        "untranslated text in English output",
                        f"Chinese character(s) found in English field '{field}': {value[:80]}",
                        "warning",
                    )
                    break  # one warning per row is enough

            intensity_value_key = find_field(row, INTENSITY_VALUE_FIELDS)
            intensity_detail_key = find_field(row, INTENSITY_DETAIL_FIELDS)
            intensity_value = normalise_value(row.get(intensity_value_key)) if intensity_value_key else ""
            intensity_detail = normalise_value(row.get(intensity_detail_key)) if intensity_detail_key else ""
            if intensity_value and intensity_value != "未找到" and "来源" not in intensity_detail:
                add(
                    path,
                    row_number,
                    "intensity source detail missing",
                    "When an intensity value is filled in Chinese output, add date/period and source in Intensity (Details).",
                    "warning",
                )

            missing_source_labels = []
            for field_name in INSTRUMENT_QUANTITATIVE_FIELDS:
                key = find_field(row, [field_name])
                value = normalise_value(row.get(key)) if key else ""
                if value and CONCRETE_DATA_PATTERN.search(value) and SOURCE_MARKER_EN not in value.lower():
                    missing_source_labels.append(key or field_name)
            if missing_source_labels:
                add(
                    path,
                    row_number,
                    "instrument quantitative source missing",
                    f"Instrument-specific/GHG fields with numeric data missing source attribution (add Source: ...): {', '.join(missing_source_labels[:20])}",
                    "warning",
                )

        activity_value = get_value(row, ["Activity", "activity"])
        for warning in _check_activity_vocabulary(activity_value):
            add(path, row_number, "activity vocabulary", warning, "warning")

        for canonical, aliases in REQUIRED_FIELD_GROUPS:
            if not get_value(row, aliases):
                add(path, row_number, "required fields missing", f"Missing required field: {canonical}")
        if not any(get_value(row, aliases) for aliases in NAME_FIELD_GROUPS):
            add(path, row_number, "required fields missing", "At least one policy name is required.")

        instrument_id = get_value(row, ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"])
        row_type_for_id = get_value(row, ROW_TYPE_FIELDS).casefold()
        parsed_id = parse_policy_id(instrument_id) if instrument_id else None
        if instrument_id and not parsed_id:
            add(
                path,
                row_number,
                "invalid instrument ID",
                "Policy Instrument ID must match {ISO3}{group code}{approach code}I{2-digit instrument sequence}S{3-digit subscheme sequence}, e.g. CHNTRARECI01S000.",
            )
        elif parsed_id:
            country_code, group_code, approach_code = expected_id_codes(
                get_value(row, ["Country", "country"]),
                get_value(row, GROUP_FIELDS),
                get_value(row, APPROACH_FIELDS),
            )
            if country_code and parsed_id["country"] != country_code:
                add(path, row_number, "invalid instrument ID", f"ID country code {parsed_id['country']} does not match row country {country_code}.")
            if group_code and parsed_id["group"] != group_code:
                add(path, row_number, "invalid instrument ID", f"ID group code {parsed_id['group']} does not match row group code {group_code}.")
            if approach_code and parsed_id["approach"] != approach_code:
                add(path, row_number, "invalid instrument ID", f"ID approach code {parsed_id['approach']} does not match row approach code {approach_code}.")
            if row_type_for_id == "instrument" and parsed_id["subscheme"] != "000":
                add(path, row_number, "invalid instrument ID", "Instrument rows must use S000.")
            if row_type_for_id in {"subscheme", "sub-scheme", "sub scheme"} and parsed_id["subscheme"] == "000":
                add(path, row_number, "invalid instrument ID", "Subscheme rows must not use S000.")

        category = get_value(row, CATEGORY_FIELDS)
        group = get_value(row, GROUP_FIELDS)
        approach = get_value(row, APPROACH_FIELDS)
        if "_cn_" in path.name:
            group_raw = get_raw_value(row, GROUP_FIELDS)
            approach_raw = get_raw_value(row, APPROACH_FIELDS)
            if group_raw and valid_cn_groups and group_raw not in valid_cn_groups:
                add(path, row_number, "invalid IFCMA category/type", f"Invalid type/group: {group_raw}")
            if approach_raw and valid_cn_approaches and approach_raw not in valid_cn_approaches:
                add(path, row_number, "invalid IFCMA category/type", f"Invalid approach: {approach_raw}")
        else:
            if category and valid_categories and category not in valid_categories:
                add(path, row_number, "invalid IFCMA category/type", f"Invalid category: {category}")
            if group:
                valid_groups = set().union(*valid_groups_by_category.values()) if valid_groups_by_category else set()
                if valid_groups and group not in valid_groups:
                    add(path, row_number, "invalid IFCMA category/type", f"Invalid type/group: {group}")
                if category and valid_groups_by_category and group not in valid_groups_by_category.get(category, set()):
                    add(path, row_number, "invalid IFCMA category/type", f"Group '{group}' is not valid for category '{category}'.")
            if approach and valid_approaches and approach not in valid_approaches:
                add(path, row_number, "invalid IFCMA category/type", f"Invalid approach: {approach}")

        economic_sector = get_value(row, ["Economic sector"])
        invalid_isic = [code for code in split_multi(economic_sector) if not is_valid_isic_division(code)]
        if invalid_isic:
            add(
                path,
                row_number,
                "invalid ISIC economic sector",
                "Economic sector must use ISIC Rev.4 section-letter plus two-digit division code(s), e.g. D35 or C29. "
                f"Invalid value(s): {', '.join(invalid_isic)}",
            )

        if normalise_control_value(approach).casefold() == RENEWABLE_REC_APPROACH:
            compliance_i = get_value(row, COMPLIANCE_I_FIELDS)
            compliance_ii = get_value(row, COMPLIANCE_II_FIELDS)
            combined_compliance = f"{compliance_i} {compliance_ii}"
            if (
                instrument_id == "CHNTRARECI02S000"
                and combined_compliance
                and not all(term in combined_compliance for term in ("物理电量", "绿证"))
                and not all(term in combined_compliance.casefold() for term in ("physical", "certificate"))
            ):
                add(
                    path,
                    row_number,
                    "renewable compliance accounting incomplete",
                    (
                        "Renewable electricity consumption responsibility-weight rows should distinguish "
                        "physical renewable electricity accounting from GEC/certificate-based accounting."
                    ),
                    "warning",
                )

        status = get_value(row, STATUS_FIELDS).lower()
        start_date = parse_date(get_value(row, START_DATE_FIELDS))
        end_date = parse_date(get_value(row, END_DATE_FIELDS))
        today = dt.date.today()
        if status and status not in VALID_STATUSES:
            add(path, row_number, "status inconsistent with start/end date", f"Invalid status: {status}")
        if start_date and end_date and end_date < start_date:
            add(path, row_number, "status inconsistent with start/end date", "End date is earlier than start date.")
        if status == "scheduled" and start_date and start_date <= today:
            add(path, row_number, "status inconsistent with start/end date", "Scheduled instrument has a start date on or before today.")
        if status == "in force" and end_date and end_date < today:
            add(path, row_number, "status inconsistent with start/end date", "In-force instrument has an end date before today.")
        if status == "ended" and not end_date:
            add(path, row_number, "status inconsistent with start/end date", "Ended instrument has no end date.")
        if status == "ended" and end_date and end_date > today:
            add(path, row_number, "status inconsistent with start/end date", "Ended instrument has an end date after today.")

        source_values = [get_value(row, [field]) for field in SOURCE_URL_FIELDS]
        for source_value in source_values:
            for url in iter_urls(source_value):
                if not official_url(url, official_suffixes):
                    add(path, row_number, "non-official source_url", f"Non-official or unrecognized source URL: {url}", "warning")
                if consultation_or_lead_url(url):
                    add(
                        path,
                        row_number,
                        "non-canonical official source_url",
                        (
                            "Consultation or lead portal URL should be replaced by the final formal ministry "
                            f"notice page when available: {url}"
                        ),
                        "warning",
                    )
                if check_urls:
                    if url not in url_health_cache:
                        url_health_cache[url] = check_url_health(url)
                    status_code, error = url_health_cache[url]
                    if status_code in (404, 410):
                        add(path, row_number, "broken source URL", f"URL returned {status_code}: {url}")
                    elif status_code == 403:
                        add(path, row_number, "restricted source URL", f"URL returned 403 (may be geo-restricted): {url}", "warning")
                    elif status_code >= 500:
                        add(path, row_number, "source URL temporarily unreachable", f"URL returned {status_code}: {url}", "warning")
                    elif status_code == 0:
                        add(path, row_number, "source URL health check failed", f"URL unreachable: {error}: {url}", "warning")

        legal_document = get_value(row, LEGAL_DOCUMENT_FIELDS)
        legal_document_urls = iter_urls(legal_document)
        if len(legal_document_urls) > 1:
            add(
                path,
                row_number,
                "multiple legal document links",
                "Legal document should contain exactly one primary legal/instrument URL; put extra URLs in Other weblinks.",
            )

        confidence = get_value(row, CONFIDENCE_FIELDS)
        if confidence:
            try:
                if float(confidence) < 0.7:
                    add(path, row_number, "confidence_score below 0.7", f"confidence_score={confidence}", "warning")
            except ValueError:
                add(path, row_number, "confidence_score below 0.7", f"Non-numeric confidence_score={confidence}", "warning")

        review_value = get_value(row, REVIEW_FIELDS)
        if truthy(review_value):
            add(path, row_number, "rows marked needs_human_review", "Row is marked needs_human_review.", "warning")

        row_type = get_value(row, ROW_TYPE_FIELDS).lower()
        if row_type in {"subscheme", "sub-scheme", "sub scheme"}:
            parent = get_value(row, PARENT_FIELDS)
            instrument_id = get_value(row, ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"])
            inferred_parent = infer_parent_id(instrument_id)
            if not parent and not inferred_parent:
                add(path, row_number, "sub-scheme without parent instrument", "Subscheme row has no parent instrument field or inferable parent ID.")
            elif parent and parent not in all_ids_by_file[path]:
                add(path, row_number, "sub-scheme without parent instrument", f"Parent instrument ID not found: {parent}")
            elif inferred_parent and inferred_parent != instrument_id and inferred_parent not in all_ids_by_file[path]:
                add(path, row_number, "sub-scheme without parent instrument", f"Inferred parent instrument ID not found: {inferred_parent}")

    for (_, instrument_id), locations in id_locations.items():
        if len(locations) > 1:
            for location in locations:
                file_name, row_number = location.rsplit(":", 1)
                add(Path(file_name), int(row_number), "duplicate instruments", f"Duplicate Policy Instrument ID: {instrument_id}")

    for (_, country, name, group), locations in duplicate_keys.items():
        if len(locations) > 1:
            detail = f"Duplicate country/name/group key: {country} / {name} / {group}"
            for location in locations:
                file_name, row_number = location.rsplit(":", 1)
                add(Path(file_name), int(row_number), "duplicate instruments", detail, "warning")

    rows_by_id = {
        (path, get_value(row, ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"])): (path, row_number, row)
        for path, row_number, row in flattened
        if get_value(row, ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"])
    }
    children_by_parent: dict[tuple[Path, str], list[tuple[Path, int, dict[str, object]]]] = defaultdict(list)
    for path, row_number, row in flattened:
        row_type = get_value(row, ROW_TYPE_FIELDS)
        instrument_id = get_value(row, ["Policy Instrument ID", "policy_instrument_id", "instrument_id", "id"])
        if row_type == "Subscheme" and instrument_id:
            parent_id = get_value(row, PARENT_FIELDS) or infer_parent_id(instrument_id)
            if parent_id:
                children_by_parent[(path, parent_id)].append((path, row_number, row))

    for (path, parent_id), children in children_by_parent.items():
        parent = rows_by_id.get((path, parent_id))
        if not parent:
            continue
        parent_path, parent_row_number, parent_row = parent
        for field_names, label in ((["Emission sector"], "emission sector"), (["Economic sector"], "economic sector")):
            parent_values = set(split_multi(get_value(parent_row, field_names)))
            child_values = {
                value
                for _, _, child_row in children
                for value in split_multi(get_value(child_row, field_names))
            }
            missing = child_values - parent_values
            if missing:
                add(
                    parent_path,
                    parent_row_number,
                    "parent/subscheme aggregate mismatch",
                    f"Parent {label} does not include subscheme value(s): {', '.join(sorted(missing))}",
                    "warning",
                )

    return issues, stats


def discover_dataset_files(outputs_dir: Path) -> list[Path]:
    supported = {".csv", ".tsv", ".json", ".jsonl", ".md", ".markdown", ".xlsx"}
    skip_names = {"validation_report.md", "CCPID_cn.xlsx", "CCPID_en.xlsx", "evidence_log.xlsx"}
    return sorted(
        path
        for path in outputs_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in supported and path.name not in skip_names
        and "evidence_log" not in path.stem.lower()
        and "temp" not in {part.lower() for part in path.parts}
    )


def write_report(
    report_path: Path,
    outputs_dir: Path,
    files: list[Path],
    rows_by_file: dict[Path, list[dict[str, object]]],
    issues: list[dict[str, str]],
    stats: Counter[str],
    read_errors: list[str],
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Validation Report",
        "",
        f"- Generated: {dt.datetime.now().isoformat(timespec='seconds')}",
        f"- Outputs directory: `{outputs_dir}`",
        f"- Files discovered: {len(files)}",
        f"- Rows checked: {stats.get('rows_checked', 0)}",
        f"- Issues found: {len(issues)}",
        "",
        "## Files",
        "",
    ]
    if files:
        lines.extend(f"- `{path}`: {len(rows_by_file.get(path, []))} rows" for path in files)
    else:
        lines.append("- No supported dataset files found.")

    if read_errors:
        lines.extend(["", "## Read Errors", ""])
        lines.extend(f"- {error}" for error in read_errors)

    lines.extend(["", "## Issue Summary", ""])
    if issues:
        for check, count in sorted(stats.items()):
            if check != "rows_checked":
                lines.append(f"- {check}: {count}")
    else:
        lines.append("- No validation issues found.")

    lines.extend(["", "## Issues", ""])
    if issues:
        lines.extend(
            [
                "| Severity | File | Row | Check | Detail |",
                "|---|---:|---:|---|---|",
            ]
        )
        for issue in issues:
            detail = issue["detail"].replace("|", "\\|")
            lines.append(
                f"| {issue['severity']} | `{issue['file']}` | {issue['row']} | {issue['check']} | {detail} |"
            )
    else:
        lines.append("- None.")

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Validate IFCMA dataset exports.")
    parser.add_argument("--outputs-dir", type=Path, default=DEFAULT_OUTPUTS_DIR)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH)
    parser.add_argument(
        "--official-domain-suffix",
        action="append",
        default=[],
        help="Official domain suffix to accept, e.g. .gov.cn. May be repeated.",
    )
    parser.add_argument(
        "--check-urls",
        action="store_true",
        default=False,
        help="Verify URLs are reachable (HEAD request). Adds network latency.",
    )
    args = parser.parse_args(argv)

    official_suffixes = DEFAULT_OFFICIAL_DOMAIN_SUFFIXES + tuple(args.official_domain_suffix)
    files = discover_dataset_files(args.outputs_dir)
    rows_by_file: dict[Path, list[dict[str, object]]] = {}
    read_errors: list[str] = []
    for path in files:
        try:
            rows_by_file[path] = read_dataset(path)
        except Exception as exc:  # noqa: BLE001 - report every file-level parse failure
            rows_by_file[path] = []
            read_errors.append(f"`{path}`: {exc}")

    issues, stats = validate_rows(rows_by_file, official_suffixes, args.check_urls)
    write_report(args.report, args.outputs_dir, files, rows_by_file, issues, stats, read_errors)
    print(f"Wrote validation report: {args.report}")
    return 1 if issues or read_errors else 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
